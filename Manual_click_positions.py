import os
import cv2
import openpyxl

# Create a workbook and add a sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Set the column headings
sheet.cell(row=1, column=1).value = 'Frame'
sheet.cell(row=1, column=2).value = 'mx'
sheet.cell(row=1, column=3).value = 'my'

# Create a mouse callback function
def on_mouse_click(event, x, y, flags, params):
    if event == cv2.EVENT_LBUTTONDOWN:
        # Mark the clicked position with a red circle
        cv2.circle(frame, (x, y), 1, (0, 0, 255), -1)
        # Update the displayed frame
        cv2.imshow("Video", frame)
        # Store the clicked position and frame number in the .xlsx file
        sheet.append([params['frame'], x, y])

# Get the name of the video file
video_file = 'v3.avi'
video_name = os.path.splitext(video_file)[0]
# Open the video file
video_capture = cv2.VideoCapture(video_file)


frame_number = 0
while True:
    # Read a frame from the video
    ret, frame = video_capture.read()

    # Check if the video has ended
    if not ret:
        break

    # Display the frame number in the top left corner of the frame
    text = 'Frame: {}'.format(frame_number)
    cv2.putText(frame, text, (10, 25), cv2.FONT_HERSHEY_SIMPLEX,
                1, (0, 255, 255), 2, cv2.LINE_AA)

    # Display the frame
    cv2.namedWindow("Video", cv2.WINDOW_NORMAL)
    cv2.setWindowProperty('Video', cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_NORMAL)
    cv2.imshow("Video", frame)

    # Set the mouse callback function to handle clicks
    cv2.setMouseCallback("Video", on_mouse_click, {'frame': frame_number})

    # Wait for the user to press a key
    if cv2.waitKey(0) & 0xFF == ord('q'):
        break

    frame_number += 1

# Release the video capture and destroy the window
video_capture.release()
cv2.destroyAllWindows()

# Create xlsx file name based on video title
xlsx_file = f"{video_name}_clicked.xlsx"
# Save the .xlsx file
workbook.save(xlsx_file)
